"""Excel exporter with professional formatting and dashboard charts."""
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.chart import PieChart, BarChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import logging

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Class to export cleaned DataFrame to styled Excel with dashboard."""
    
    def __init__(self):
        self.template_styles()
    
    def template_styles(self) -> dict:
        """Define professional styles."""
        return {
            'header': NamedStyle(name="header", font=Font(bold=True, color="FFFFFF", size=12),
                                 fill=PatternFill(start_color="367FA9", end_color="367FA9", fill_type="solid"),
                                 alignment=Alignment(horizontal="center", vertical="center")),
            'body': NamedStyle(name="body", font=Font(size=10),
                               alignment=Alignment(horizontal="left", vertical="top", wrap_text=True)),
            'numeric': NamedStyle(name="numeric", font=Font(size=10), number_format="#,##0.00"),
        }
    
    def export(self, df: pd.DataFrame, output_path: str) -> None:
        """Export DataFrame to Excel with Data sheet and Dashboard."""
        wb = openpyxl.Workbook()
        
        # Sheet 1: Data
        ws_data = wb.active
        ws_data.title = "Data"
        self._write_df_to_sheet(ws_data, df)
        self._style_data_sheet(ws_data)
        
        # Sheet 2: Dashboard
        ws_dash = wb.create_sheet("Dashboard")
        self._create_dashboard(ws_dash, df)
        
        wb.save(output_path)
        logger.info(f"Excel exported to {output_path}")
    
    def _write_df_to_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame) -> None:
        """Write DataFrame to worksheet."""
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    
    def _style_data_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Apply professional styling to Data sheet."""
        # Headers
        for cell in ws[1]:
            cell.style = self.template_styles()['header']
        
        # Body
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter in ['A', 'B', 'C']:  # Text columns example
                    cell.style = self.template_styles()['body']
                else:
                    cell.style = self.template_styles()['numeric']
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header
        ws.freeze_panes = 'A2'
    
    def _create_dashboard(self, ws: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame) -> None:
        """Create dashboard charts."""
        # Title
        ws['A1'] = "Seeker Logs Dashboard"
        ws['A1'].font = Font(bold=True, size=16)
        
        # OS Pie Chart (starting row 3)
        os_count = df['OS'].value_counts().head(10).reset_index()
        os_count.columns = ['OS', 'Count']
        for r_idx, row in enumerate(dataframe_to_rows(os_count, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=3, max_row=3+len(os_count))
        data = Reference(ws, min_col=2, min_row=2, max_row=2+len(os_count))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distribución de OS"
        ws.add_chart(pie, "D3")
        
        # ISP Bar Chart (starting row 15)
        isp_count = df['ISP'].value_counts().head(10).reset_index()
        isp_count.columns = ['ISP', 'Count']
        start_row = 15
        for r_idx, row in enumerate(dataframe_to_rows(isp_count, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        bar = BarChart()
        data = Reference(ws, min_col=2, min_row=start_row-1, max_row=start_row-1+len(isp_count))
        cats = Reference(ws, min_col=1, min_row=start_row, max_row=start_row+len(isp_count))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.title = "Top 10 ISP"
        ws.add_chart(bar, "D15")
        
        # Top 10 Cities (starting row 30)
        cities_count = df['City'].value_counts().head(10).reset_index()
        cities_count.columns = ['City', 'Count']
        start_row_city = 30
        for r_idx, row in enumerate(dataframe_to_rows(cities_count, index=False, header=True), start_row_city):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        bar_city = BarChart()
        data_city = Reference(ws, min_col=2, min_row=start_row_city-1, max_row=start_row_city-1+len(cities_count))
        cats_city = Reference(ws, min_col=1, min_row=start_row_city, max_row=start_row_city+len(cities_count))
        bar_city.add_data(data_city, titles_from_data=True)
        bar_city.set_categories(cats_city)
        bar_city.title = "Top 10 Ciudades"
        ws.add_chart(bar_city, "D30")
